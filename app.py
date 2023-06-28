from flask import Flask, render_template, request, redirect, url_for
import json
from datetime import datetime
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
import webview

app = Flask(__name__)
archivo_excel = 'productos.xlsx'
embalaje_agregado = False


# Cargar productos desde el archivo JSON
def cargar_productos():
    try:
        workbook = load_workbook(archivo_excel)
        sheet = workbook.active
        
        productos = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            producto = {
                'codigo': row[0],
                'descripcion': row[1],
                'precio_minorista': row[2],
                'precio_mayorista': row[3],
                'cantidad': row[4],
            }
            productos.append(producto)
        
        return productos
    except FileNotFoundError:
        return []


# Guardar productos en la planilla de Excel
def guardar_productos(productos):
    workbook = load_workbook(archivo_excel)
    sheet = workbook.active
    
    # Borrar contenido existente en la hoja
    sheet.delete_rows(2, sheet.max_row)
    
    # Escribir los productos en la hoja
    for i, producto in enumerate(productos, start=2):
        for j, value in enumerate(producto.values(), start=1):
            sheet[f'{get_column_letter(j)}{i}'] = value
    
    # Guardar los cambios en el archivo
    workbook.save(archivo_excel)


# Página principal
@app.route('/')
def index():
    return render_template('index.html')


# Página de inventario
@app.route('/inventario', methods=['GET', 'POST'])
def inventario():
    if request.method == 'POST':
        codigo = request.form['codigo']
        descripcion = request.form['descripcion']
        precio_minorista = request.form['precio_minorista']
        precio_mayorista = request.form['precio_mayorista']
        cantidad = int(request.form['cantidad'])
        
        productos = cargar_productos()
        
        productos.append({
            'codigo': codigo,
            'descripcion': descripcion,
            'precio_minorista': precio_minorista,
            'precio_mayorista': precio_mayorista,
            'cantidad': cantidad
        })
        
        guardar_productos(productos)
        
        return redirect(url_for('inventario'))
    
    productos = cargar_productos()
    return render_template('inventario.html', productos=productos)


# Ruta para realizar un ingreso
@app.route('/ingreso', methods=['GET', 'POST'])
def ingreso():
    if request.method == 'POST':
        codigo = int(request.form['codigo'])
        cantidad = int(request.form['cantidad'])
        
        productos = cargar_productos()
        
        # Buscar el producto por su código y actualizar la cantidad en el inventario
        for producto in productos:
            if producto['codigo'] == codigo:
                producto['cantidad'] += cantidad
                break
        
        guardar_productos(productos)
        
        return render_template('ingreso.html')
    
    return render_template('ingreso.html')


# Ruta para agregar un producto al inventario
@app.route('/agregar_producto', methods=['GET', 'POST'])
def agregar_producto():
    if request.method == 'POST':
        codigo = int(request.form['codigo'])
        descripcion = request.form['descripcion']
        cantidad = int(request.form['cantidad'])
        precio_minorista = int(request.form['precio_minorista'])
        precio_mayorista = int(request.form['precio_mayorista'])
        
        productos = cargar_productos()
        
        productos.append({
            'codigo': codigo,
            'descripcion': descripcion,
            'precio_minorista': precio_minorista,
            'precio_mayorista': precio_mayorista,
            'cantidad': cantidad
        })
        
        guardar_productos(productos)
        
        return redirect(url_for('inventario'))
    
    return render_template('agregar_producto.html')


def guardar_carrito_en_archivo(carrito):
    with open('carrito.json', 'w') as file:
        json.dump(carrito, file)


# Guarda el carrito de compras en las cookies
def guardar_carrito(carrito):
    guardar_carrito_en_archivo(carrito)
    return redirect(url_for('venta'))


@app.route('/limpiar_carrito', methods=['POST'])
# Limpiar carrito de compras
def limpiar_carrito():
    global embalaje_agregado
    embalaje_agregado = False
    guardar_carrito_en_archivo([])
    return redirect(url_for('venta'))


# Carga el carrito de compras desde el archivo json
def cargar_carrito():
    try:
        with open('carrito.json', 'r') as file:
            carrito = json.load(file)
    except (FileNotFoundError, json.JSONDecodeError):
        carrito = []
    return carrito


# Ruta para realizar una venta
@app.route('/venta', methods=['GET', 'POST'])
def venta():
    if request.method == 'POST':
        descripcion = request.form['descripcion']
        cantidad = int(request.form['cantidad'])
        
        # Obtén el tipo de precio seleccionado
        tipo_precio = request.form['tipo_precio']
        
        # Obtén el tipo de pago seleccionado
        tipo_pago = request.form['tipo_pago']
        
        productos = cargar_productos()
        
        # Buscar el producto por su código
        producto_encontrado = None
        for producto in productos:
            if producto['descripcion'] == descripcion:
                producto_encontrado = producto
                break
        
        if producto_encontrado and producto_encontrado['cantidad'] >= cantidad:
            if tipo_precio == 'Minorista' and tipo_pago == 'Efectivo':
                precio = int(producto_encontrado['precio_minorista'])
            elif tipo_precio == 'Mayorista' and tipo_pago == 'Efectivo':
                precio = int(producto_encontrado['precio_mayorista'])
            elif tipo_precio == 'Minorista' and tipo_pago == 'Transf':
                precio = (int(producto_encontrado['precio_minorista'])) + \
                         (int(producto_encontrado['precio_minorista']) * 0.05)
            elif tipo_precio == 'Mayorista' and tipo_pago == 'Transf':
                precio = (int(producto_encontrado['precio_mayorista'])) + \
                         (int(producto_encontrado['precio_mayorista']) * 0.05)
            
            else:
                return 'Tipo de precio inválido'
            
            # Calcula el subtotal
            subtotal = precio * cantidad
            
            # Agrega el producto al carrito de compras
            carrito = cargar_carrito()
            carrito.append({'codigo': producto_encontrado['codigo'],
                            'cantidad': cantidad,
                            'descripcion': producto_encontrado['descripcion'],
                            'subtotal': subtotal,
                            'tipo_precio': tipo_precio,
                            'tipo_pago': tipo_pago})
            return guardar_carrito(carrito)
        else:
            return 'Producto no disponible o cantidad insuficiente'
    
    productos = cargar_productos()
    carrito = cargar_carrito()
    
    # Obtén la lista de descripciones de los productos
    descripciones_productos = [producto['descripcion'] for producto in productos]
    
    # Suma el subtotal al total de la venta
    total_venta_carrito = 0
    for item in carrito:
        total_venta_carrito += item['subtotal']
    
    # Agrega la variable 'embalaje' al contexto
    embalaje = False
    
    return render_template('venta.html', productos=productos, carrito=carrito,
                           descripciones_productos=descripciones_productos, total_venta_carrito=total_venta_carrito,
                           embalaje=embalaje)


# Ruta para aplicar el embalaje
@app.route('/aplicar_embalaje', methods=['POST'])
def aplicar_embalaje():
    global embalaje_agregado
    embalaje_agregado = True
    
    # Obtiene el carrito de compras y el total de la venta actual
    carrito = cargar_carrito()
    total_venta_carrito = sum(item['subtotal'] for item in carrito)
    
    # Agrega el precio del embalaje (700) al total de la venta
    total_venta_carrito += 700
    
    # Actualiza el carrito con el nuevo total de la venta
    for item in carrito:
        item['total_venta'] = total_venta_carrito
    
    # Renderiza la plantilla de venta con el carrito actualizado y el nuevo total de la venta
    return render_template('venta.html', productos=cargar_productos(), carrito=carrito,
                           descripciones_productos=[producto['descripcion'] for producto in cargar_productos()],
                           total_venta_carrito=total_venta_carrito)


# Ruta para imprimir el ticket y marcarlo como impreso en el archivo Excel
@app.route('/imprimir_ticket', methods=['GET', 'POST'])
def imprimir_ticket():
    carrito = cargar_carrito()
    productos = cargar_productos()
    total_venta = 0
    now = datetime.now()
    
    # Actualiza el inventario y calcula el total de la venta
    for item in carrito:
        codigo = item['codigo']
        cantidad = item['cantidad']
        
        producto_encontrado = None
        for producto in productos:
            if producto['codigo'] == codigo:
                producto_encontrado = producto
                break
        
        if producto_encontrado and producto_encontrado['cantidad'] >= cantidad:
            # Actualiza la cantidad del producto en el inventario
            producto_encontrado['cantidad'] -= cantidad
            
            # Calcula el subtotal y agrega al total de la venta
            subtotal = item['subtotal']
            total_venta += subtotal
        else:
            return 'Producto no disponible o cantidad insuficiente'
    
    # Aplica el precio del embalaje si está seleccionado
    # Si el embalaje fue agregado, se agrega el precio del embalaje (700) al total de la venta
    if embalaje_agregado:
        total_venta += 700
    else:
        total_venta = total_venta
    
    # Guarda los cambios en el inventario
    guardar_productos(productos)
    
    # Elimina el contenido del archivo carrito.json
    guardar_carrito_en_archivo([])
    
    # Registra la venta en el archivo de registro Excel
    guardar_venta_en_excel(carrito, total_venta)
    
    ticket_html = render_template('ticket.html', ventas=carrito, total_venta=total_venta, now=now,
                                  embalaje_agregado=embalaje_agregado)
    
    return ticket_html


def guardar_venta_en_excel(ventas, total_venta):
    # Cargar el archivo existente o crear uno nuevo si no existe
    try:
        workbook = load_workbook('registro_ventas.xlsx')
    except FileNotFoundError:
        workbook = Workbook()
    
    # Obtener la primera hoja (detalles de las ventas)
    sheet = workbook.active
    
    # Obtener la segunda hoja (total de ventas)
    sheet_total = workbook['Total Ventas']
    
    # Si la hoja 'Total Ventas' no existe, crearla
    if sheet_total is None:
        sheet_total = workbook.create_sheet('Total Ventas')
    
    # Obtener el número de filas existentes en la primera hoja
    last_row = sheet.max_row
    
    # Obtener la fecha y hora actual
    now = datetime.now()
    fecha_hora = now.strftime('%Y-%m-%d %H:%M:%S')
    
    # Agregar las ventas en la primera hoja
    for venta in ventas:
        descripcion = venta['descripcion']
        cantidad = venta['cantidad']
        tipo_pago = venta['tipo_pago']
        subtotal = venta['subtotal']
        
        # Aumentar el número de filas
        last_row += 1
        
        # Escribir los datos en las columnas correspondientes
        sheet['A{}'.format(last_row)] = fecha_hora
        sheet['B{}'.format(last_row)] = descripcion
        sheet['C{}'.format(last_row)] = cantidad
        sheet['D{}'.format(last_row)] = tipo_pago
        sheet['E{}'.format(last_row)] = subtotal
    
    # Agregar la fecha y el monto total en la segunda hoja
    last_row_total = sheet_total.max_row + 1
    sheet_total['A{}'.format(last_row_total)] = fecha_hora
    sheet_total['B{}'.format(last_row_total)] = total_venta
    
    # Guardar los cambios en el archivo
    workbook.save('registro_ventas.xlsx')


if __name__ == '__main__':
    webview.create_window("WOSSA-STOCK", app)
    webview.start()
